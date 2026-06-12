"""
Excel → CSV 配表导出工具（支持多 Sheet 导出）

使用方法：
  1. python export_config.py --template              # 生成空白 Excel 模板（含 3 个 Sheet）
  2. 在 Excel 中编辑保存为 .xlsx
  3. python export_config.py                          # 导出所有 Sheet 为独立 CSV
  4. python export_config.py --sheet SkillConfig      # 仅导出指定 Sheet

依赖安装：
  pip install openpyxl

修改规范：新增字段只需：
  1. 在 Excel 中添加一列（第一行写英文名）
  2. 在对应的 C# Config 类中添加同名 public 字段
"""

import argparse
import csv
import os
from openpyxl import load_workbook, Workbook


def excel_to_csv(input_path, output_path, sheet_name):
    """将 Excel 的指定工作表导出为 CSV 配表"""
    wb = load_workbook(input_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"错误：工作表 '{sheet_name}' 不存在")
        return False

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"错误：工作表 '{sheet_name}' 为空")
        return False

    # 第1行作为 CSV 表头，原样保留
    headers = [str(cell).strip() if cell else "" for cell in rows[0]]
    headers = [h for h in headers if h]

    if not headers:
        print(f"错误：工作表 '{sheet_name}' 第一行为空")
        return False

    print(f"\n[{sheet_name}] 表头: {headers}")

    csv_rows = [headers]
    last_vals = [""] * len(headers)
    for row_idx, row in enumerate(rows[1:], start=2):
        values = []
        for col_idx in range(len(headers)):
            cell_val = row[col_idx] if col_idx < len(row) else None
            # 合并单元格：空单元格继承上一行同一列的值
            if cell_val is None or (isinstance(cell_val, str) and not cell_val.strip()):
                cell_val = last_vals[col_idx]
            elif isinstance(cell_val, float):
                cell_val = str(int(cell_val)) if cell_val == int(cell_val) else str(cell_val)
            else:
                cell_val = str(cell_val).strip()
            last_vals[col_idx] = cell_val
            values.append(cell_val)

        id_val = values[0]
        if not id_val:
            print(f"  [{sheet_name}] 跳过第 {row_idx} 行：Id 为空")
            continue

        csv_rows.append(values)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(csv_rows)

    print(f"  → 导出 {len(csv_rows) - 1} 条数据 -> {output_path}")
    return True


def export_all_sheets(input_path, output_dir):
    """导出所有工作表为独立的 CSV 文件"""
    wb = load_workbook(input_path, data_only=True)
    success_count = 0
    for sheet_name in wb.sheetnames:
        output_path = os.path.join(output_dir, f"{sheet_name}.csv")
        if excel_to_csv(input_path, output_path, sheet_name):
            success_count += 1
    print(f"\n== 全部导出完成: {success_count}/{len(wb.sheetnames)} 个工作表")


def list_sheets(input_path):
    """列出 Excel 文件中的所有工作表"""
    wb = load_workbook(input_path, read_only=True)
    print("\n工作表列表：")
    for name in wb.sheetnames:
        print(f"  - {name}")


def create_template(output_path):
    """生成空白 Excel 模板文件（3 个 Sheet）"""

    # ==================== Sheet 1: 角色配置 ====================
    char_columns = [
        ("Id",            1001,       "角色ID"),
        ("Name",          "日和莉",   "中文名"),
        ("OriginName",    "Hiyori",   "名称"),
        ("SpineId",       "spine_1001", "Spine地址"),
        ("HP",            5800,       "生命值"),
        ("PhysicalAttack",1580,       "物理攻击力"),
        ("MagicAttack",   0,          "魔法攻击力"),
        ("AttackRange",   200,        "攻击范围"),
        ("AnimRunGameStart", "01_run_gamestart", "开局跑步动画"),
        ("AnimStandBy",   "01_standBy",  "准备动画"),
        ("AnimRun",       "01_run",      "跑步动画"),
        ("AnimIdle",      "01_multy_idle_standBy", "待机动画"),
        ("AttackId",      "100101_attack", "普攻ID"),
        ("UbSkillId",     "100101_skill0", "UB技能ID"),
        ("Skill1Id",      "100101_skill1", "1技能ID"),
        ("Skill2Id",      "100101_skill2", "2技能ID"),
        ("StartSequence", "21",          "启动序列"),
        ("LoopSequence",  "AA2A1",       "循环序列"),
    ]

    # ==================== Sheet 2: 技能配置 ====================
    skill_columns = [
        ("CharacterId", "",         "角色ID"),
        ("Id",        100101,     "技能ID"),
        ("Name",      "烈焰斩",   "技能名称"),
        ("CastTime",  1.5,        "前摇时长"),
        ("AnimName",  "skill_1",  "技能动画名"),
        ("SoundName", "skill_ub", "技能音效名"),
    ]

    # ==================== Sheet 3: 技能效果 ====================
    effect_columns = [
        ("SkillId",      100101,     "所属技能ID"),
        ("EffectIndex",  1,          "效果序号(按顺序执行)"),
        ("EffectType",   "Damage",   "效果类型(Damage/Heal/Buff/Debuff)"),
        ("EffectTarget", "SingleEnemy", "目标(SingleEnemy/AllEnemies/Self/AllAllies)"),
        ("DamageType",   "Physical", "伤害类型(Physical/Magic)"),
        ("SkillMulti",   3.0,        "攻击力倍率"),
        ("EffectValue",  0,          "基础数值"),
        ("SkillLevelMulti", 0,       "技能等级倍率"),
        ("CastFrame",    50,         "效果前摇(帧)"),
    ]

    wb = Workbook()

    # ---- Sheet 1: 角色配置 ----
    ws1 = wb.active
    ws1.title = "CharacterConfig"
    ws1.append([col[0] for col in char_columns])
    ws1.append([col[2] for col in char_columns])
    ws1.append([col[1] for col in char_columns])
    for i in range(len(char_columns)):
        ws1.column_dimensions[chr(65 + i)].width = 16
    ws1.freeze_panes = "A3"

    # ---- Sheet 2: 技能配置 ----
    ws2 = wb.create_sheet("SkillConfig")
    ws2.append([col[0] for col in skill_columns])
    ws2.append([col[2] for col in skill_columns])
    ws2.append([col[1] for col in skill_columns])
    for i in range(len(skill_columns)):
        ws2.column_dimensions[chr(65 + i)].width = 14
    ws2.freeze_panes = "A3"

    # ---- Sheet 3: 技能效果 ----
    ws3 = wb.create_sheet("SkillEffectConfig")
    ws3.append([col[0] for col in effect_columns])
    ws3.append([col[2] for col in effect_columns])
    ws3.append([col[1] for col in effect_columns])
    for i in range(len(effect_columns)):
        ws3.column_dimensions[chr(65 + i)].width = 16
    ws3.freeze_panes = "A3"

    # 保存
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    print(f"\n== 模板已生成: {output_path}")
    print(f"   包含 3 个工作表: CharacterConfig / SkillConfig / SkillEffectConfig")
    print(f"   每表第1行=英文字段名，第2行=中文注释，第3行=示例数据")
    print(f"   编辑完成后运行 python export_config.py 导出 CSV")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel → CSV 配表导出工具")
    parser.add_argument("--input", default=None,
                        help="输入的 Excel 文件路径（默认：../Assets/Configs/CharacterConfig.xlsx）")
    parser.add_argument("--output", default=None,
                        help="输出的 CSV 文件路径（默认：../Assets/Configs/ 目录下同名文件）")
    parser.add_argument("--sheet", default=None,
                        help="指定导出某个工作表（可选，默认导出所有工作表）")
    parser.add_argument("--template", action="store_true",
                        help="生成空白 Excel 模板文件并退出")
    parser.add_argument("--list-sheets", action="store_true",
                        help="列出 Excel 文件中的所有工作表并退出")

    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_path = args.input or os.path.join(script_dir, "..", "Assets", "Configs", "CharacterConfig.xlsx")
    output_path = args.output or os.path.join(script_dir, "..", "Assets", "Configs", "CharacterConfig.csv")

    input_path = os.path.abspath(input_path)
    output_path = os.path.abspath(output_path)

    if args.list_sheets:
        if not os.path.exists(input_path):
            print(f"文件不存在：{input_path}")
        else:
            list_sheets(input_path)
    elif args.template:
        create_template(input_path)
    else:
        if not os.path.exists(input_path):
            print(f"❌ 文件不存在：{input_path}")
            print("请先在 Excel 中创建配表并保存为 .xlsx 格式")
        else:
            if args.sheet:
                excel_to_csv(input_path, output_path, args.sheet)
            else:
                output_dir = os.path.dirname(output_path)
                export_all_sheets(input_path, output_dir)
